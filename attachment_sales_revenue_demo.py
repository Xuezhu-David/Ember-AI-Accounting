"""Instruction plus Excel attachment demo for AI voucher generation (LLM-powered).

Uses an LLM to generate vouchers from Excel attachment data.

Run:
    source .venv/bin/activate
    python attachment_sales_revenue_demo.py
"""

import asyncio
import json
from pathlib import Path

from agentscope.message import Msg
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from attachment_voucher_agent import AttachmentVoucherAgent
from llm_voucher_generator import MODEL_NAME

INPUT_PATH = Path("data/input/sales_revenue_attachment.xlsx")


def create_sample_sales_excel(path: Path) -> None:
    """Create a sample Excel attachment similar to a business export."""

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售收入明细"

    headers = [
        "业务流水号",
        "公司代码",
        "凭证日期",
        "过账日期",
        "客户编码",
        "客户名称",
        "产品类型",
        "合同号",
        "发票号",
        "币种",
        "税率",
        "不含税金额",
        "税额",
        "价税合计",
        "利润中心",
        "成本中心",
    ]
    rows = [
        [
            "SO-20260520-001",
            "1000",
            "2026-05-20",
            "2026-05-20",
            "C10086",
            "上海演示客户有限公司",
            "software",
            "CTR-2026-SW-001",
            "INV-20260520-0001",
            "CNY",
            0.13,
            100000.00,
            13000.00,
            113000.00,
            "PC-SOFTWARE",
            "CC-SALES",
        ],
        [
            "SO-20260520-002",
            "1000",
            "2026-05-20",
            "2026-05-20",
            "C10087",
            "北京样例科技有限公司",
            "service",
            "CTR-2026-SV-002",
            "INV-20260520-0002",
            "CNY",
            0.06,
            50000.00,
            3000.00,
            53000.00,
            "PC-SERVICE",
            "CC-SALES",
        ],
    ]

    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 12),
            28,
        )

    table = Table(displayName="SalesRevenueInput", ref=f"A1:P{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A2"
    workbook.save(path)


async def main() -> None:
    create_sample_sales_excel(INPUT_PATH)

    request = {
        "instruction": "请根据附件中的销售收入明细，生成 SAP 销售收入凭证草稿。",
        "attachments": [str(INPUT_PATH)],
    }
    msg = Msg(
        name="finance_user",
        role="user",
        content=json.dumps(request, ensure_ascii=False),
    )

    agent = AttachmentVoucherAgent("attachment_voucher_agent")
    print("User instruction and attachment:")
    print(json.dumps(request, ensure_ascii=False, indent=2))
    print()

    response = await agent.reply(msg)
    print()
    print("Agent response (LLM generated):")
    print(response.get_text_content())
    print()
    print(f"LLM model used: {MODEL_NAME}")


if __name__ == "__main__":
    asyncio.run(main())
