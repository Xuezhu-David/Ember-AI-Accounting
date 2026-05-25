"""Load sales revenue transactions from Excel attachments."""

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from voucher_models import SalesTransaction


HEADER_ALIASES = {
    "transaction_id": {"transaction_id", "业务流水号", "业务编号", "订单号"},
    "company_code": {"company_code", "公司代码", "公司编码"},
    "document_date": {"document_date", "凭证日期", "单据日期", "开票日期"},
    "posting_date": {"posting_date", "过账日期", "入账日期"},
    "customer_code": {"customer_code", "客户编码", "客户代码"},
    "customer_name": {"customer_name", "客户名称", "购买方名称"},
    "product_type": {"product_type", "产品类型", "收入类型", "业务类型"},
    "contract_no": {"contract_no", "合同号", "合同编号"},
    "invoice_no": {"invoice_no", "发票号", "发票号码"},
    "currency": {"currency", "币种"},
    "tax_rate": {"tax_rate", "税率"},
    "tax_excluded_amount": {"tax_excluded_amount", "不含税金额", "销售额"},
    "tax_amount": {"tax_amount", "税额", "销项税额"},
    "total_amount": {"total_amount", "价税合计", "含税金额"},
    "profit_center": {"profit_center", "利润中心"},
    "cost_center": {"cost_center", "成本中心"},
}


REQUIRED_FIELDS = set(HEADER_ALIASES)


def load_sales_transactions(path: str | Path, sheet_name: str | None = None) -> list[SalesTransaction]:
    """Read sales transactions from an Excel workbook."""

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    field_by_column = _map_headers(header_row)
    missing = sorted(REQUIRED_FIELDS - set(field_by_column.values()))
    if missing:
        raise ValueError(f"Missing required Excel columns: {', '.join(missing)}")

    transactions: list[SalesTransaction] = []
    for row_no, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue

        payload: dict[str, Any] = {}
        for column_index, cell_value in enumerate(row):
            field = field_by_column.get(column_index)
            if field:
                payload[field] = cell_value

        try:
            transactions.append(_build_transaction(payload))
        except Exception as exc:
            raise ValueError(f"Invalid sales transaction at row {row_no}: {exc}") from exc

    return transactions


def _map_headers(header_row: tuple[Any, ...]) -> dict[int, str]:
    normalized_to_field = {
        _normalize(alias): field
        for field, aliases in HEADER_ALIASES.items()
        for alias in aliases
    }

    field_by_column: dict[int, str] = {}
    for index, header in enumerate(header_row):
        if header is None:
            continue
        field = normalized_to_field.get(_normalize(str(header)))
        if field:
            field_by_column[index] = field
    return field_by_column


def _normalize(value: str) -> str:
    return value.strip().lower().replace(" ", "").replace("_", "")


def _decimal(value: Any) -> Decimal:
    if value is None or value == "":
        raise ValueError("money or rate field is empty")
    return Decimal(str(value)).quantize(Decimal("0.000001")).normalize()


def _string(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def _build_transaction(payload: dict[str, Any]) -> SalesTransaction:
    return SalesTransaction(
        transaction_id=_string(payload["transaction_id"]),
        company_code=_string(payload["company_code"]),
        document_date=_string(payload["document_date"]),
        posting_date=_string(payload["posting_date"]),
        customer_code=_string(payload["customer_code"]),
        customer_name=_string(payload["customer_name"]),
        product_type=_string(payload["product_type"]),
        contract_no=_string(payload["contract_no"]),
        invoice_no=_string(payload["invoice_no"]),
        currency=_string(payload["currency"], "CNY"),
        tax_rate=_decimal(payload["tax_rate"]),
        tax_excluded_amount=_decimal(payload["tax_excluded_amount"]),
        tax_amount=_decimal(payload["tax_amount"]),
        total_amount=_decimal(payload["total_amount"]),
        profit_center=_string(payload["profit_center"]),
        cost_center=_string(payload["cost_center"]),
    )
